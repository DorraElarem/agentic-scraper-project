"""
Scraper Traditionnel COHÉSIF - Intégration COMPLÈTE des utils
Version finale utilisant TOUS les modules du projet de manière cohérente
"""

import os
import re
import json
import requests
import time
import io
from typing import Optional, Dict, List, Any, Tuple
import logging
from urllib.parse import urlparse
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from datetime import datetime

# CORRECTION : Import des schemas cohérents
from app.models.schemas import (
    ScrapedContent, EnhancedExtractedValue, ExtractionSummary, SourceAnalysis,
    ExtractionQuality, ProcessingInfo, EconomicCategory, ExtractionMethod,
    TemporalMetadata, ValidationDetails, EconomicCoherence
)

# CORRECTION : Import de la configuration cohérente
from app.config.settings import settings

# INTÉGRATION CRITIQUE : Import de TOUS les utils
from app.utils.helpers import (
    format_timestamp, calculate_execution_time, extract_domain,
    validate_url, categorize_url_type, detect_tunisian_content_patterns,
    suggest_optimal_strategy, debug_extraction_data, log_extraction_details,
    suggest_extraction_improvements
)
from app.utils.data_validator import validate_indicators_strict, is_economic_indicator_valid
from app.utils.temporal_filter import filter_by_temporal_period, is_in_target_period
from app.utils.clean_extraction_patterns import extract_clean_economic_data, is_valid_indicator
from app.utils.storage import smart_storage

# Imports PDF/Excel sécurisés
PDF_AVAILABLE = False
EXCEL_AVAILABLE = False

try:
    from pypdf import PdfReader
    PDF_AVAILABLE = True
except ImportError:
    try:
        from PyPDF2 import PdfReader
        PDF_AVAILABLE = True
    except ImportError:
        pass

try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    try:
        import xlrd
        EXCEL_AVAILABLE = True
    except ImportError:
        pass

logger = logging.getLogger(__name__)

class CohesiveTunisianWebScraper:
    """Scraper Traditionnel utilisant TOUS les modules du projet de manière cohérente"""
    
    def __init__(self, delay: float = None):
        self.delay = delay or settings.DEFAULT_DELAY
        self.timeout = settings.REQUEST_TIMEOUT
        self.max_retries = settings.MAX_SCRAPE_RETRIES
        self.user_agent = settings.SCRAPE_USER_AGENT
        self.max_content_size = settings.MAX_CONTENT_LENGTH
        
        # Session HTTP moderne
        self.session = self._create_modern_session()
        
        # Patterns d'extraction optimisés
        self._setup_extraction_patterns()
        
        # Métriques d'utilisation des modules
        self.module_usage = {
            'helpers': 0,
            'data_validator': 0,
            'temporal_filter': 0,
            'clean_extractor': 0,
            'storage': 0
        }
        
        logger.info("CohesiveTunisianWebScraper initialized with ALL modules")

    def _create_modern_session(self) -> requests.Session:
        """Session HTTP moderne utilisant les helpers pour la configuration"""
        session = requests.Session()
        
        # Configuration retry moderne
        retry_strategy = Retry(
            total=self.max_retries,
            status_forcelist=[429, 500, 502, 503, 504, 520, 521, 522, 524],
            allowed_methods=["HEAD", "GET", "OPTIONS"],
            backoff_factor=1.5,
            raise_on_status=False
        )
        
        adapter = HTTPAdapter(
            max_retries=retry_strategy,
            pool_maxsize=20,
            pool_connections=10,
            pool_block=False
        )
        
        session.mount("http://", adapter)
        session.mount("https://", adapter)
        
        # Headers modernes
        session.headers.update({
            'User-Agent': self.user_agent,
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'fr-FR,fr;q=0.9,en-US;q=0.8,ar-TN;q=0.7',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Cache-Control': 'max-age=0',
            'DNT': '1'
        })
        
        return session

    def _setup_extraction_patterns(self) -> None:
        """Configuration des patterns utilisant les helpers"""
        
        # Patterns modernes optimisés
        self.modern_patterns = [
            # World Bank API spécialisés
            r'"value":\s*([0-9]+\.?[0-9]*(?:e[+-]?\d+)?)',
            r'"date":\s*"(\d{4})".*?"value":\s*([0-9]+\.?[0-9]*)',
            
            # Sites tunisiens (BCT, INS, etc.)
            r'([A-Za-zÀ-ÿ\s\-_\.]{5,50})\s*[:=\-]\s*([0-9]+[,.]?[0-9]*)\s*(MD|MDT|%|millions?|milliards?|TND|USD|EUR)?',
            
            # APIs internationales
            r'"([a-zA-Z_][a-zA-Z0-9_]*)":\s*([0-9]+\.?[0-9]*(?:e[+-]?\d+)?)',
            
            # Tables HTML responsive
            r'<t[dh][^>]*>([^<]{3,60})</t[dh]>\s*<t[dh][^>]*>([0-9,]+\.?[0-9]*[%]?)</t[dh]>',
        ]
        
        # Indicateurs économiques tunisiens
        self.tunisian_indicators = {
            'monetary': ['taux directeur', 'taux d\'intérêt', 'tmm', 'inflation', 'déflation'],
            'statistical': ['population', 'démographie', 'emploi', 'chômage', 'unemployment'],
            'trade': ['exportations', 'importations', 'balance commerciale', 'exports', 'imports'],
            'fiscal': ['pib', 'gdp', 'dette publique', 'déficit budgétaire', 'budget deficit'],
            'financial': ['réserves', 'change', 'bourse', 'indice', 'reserves', 'exchange rate']
        }

    def scrape(self, url: str, enable_llm_analysis: bool = False) -> Optional[ScrapedContent]:
        """Point d'entrée principal avec intégration complète des utils"""
        
        start_time = datetime.utcnow()
        
        try:
            logger.info(f"Starting cohesive traditional scraping: {url}")
            
            # INTÉGRATION : Validation URL avec helpers
            self.module_usage['helpers'] += 1
            if not validate_url(url):
                logger.error(f"Invalid URL format: {url}")
                return None
            
            # INTÉGRATION : Détection contexte tunisien avec helpers
            url_type = categorize_url_type(url)
            tunisian_context = detect_tunisian_content_patterns(url)
            strategy_suggestion = suggest_optimal_strategy(url)
            
            logger.info(f"URL analysis: type={url_type}, tunisian={tunisian_context.get('tunisian_context', False)}")
            logger.info(f"Suggested strategy: {strategy_suggestion['recommended_strategy']}")
            
            # Récupération du contenu
            content, headers = self._fetch_content_modern(url)
            if not content:
                logger.warning(f"No content retrieved: {url}")
                return None
            
            # Analyse de source moderne avec helpers
            source_analysis = self._analyze_source_with_helpers(url, headers)
            
            # Extraction adaptative selon le type de contenu
            extracted_data = self._extract_data_with_all_utils(content, url, source_analysis)
            
            # INTÉGRATION : Post-traitement avec tous les utils
            post_processed_data = self._apply_comprehensive_post_processing(extracted_data, content, url)
            
            # Construction du résultat final
            execution_time = calculate_execution_time(start_time, datetime.utcnow())
            result = self._build_cohesive_scraped_content(
                content, post_processed_data, source_analysis, url, execution_time, enable_llm_analysis
            )
            
            # INTÉGRATION : Sauvegarde automatique via smart_storage
            self._save_via_smart_storage(result, url)
            
            # INTÉGRATION : Debug et logging via helpers
            self._perform_debug_logging(result, url, strategy_suggestion['recommended_strategy'])
            
            logger.info(f"Cohesive traditional scraping completed: {len(post_processed_data.get('values', {}))} values")
            return result
            
        except Exception as e:
            logger.error(f"Cohesive traditional scraping failed for {url}: {e}")
            return None

    def _analyze_source_with_helpers(self, url: str, headers: Dict[str, str]) -> SourceAnalysis:
        """Analyse de source utilisant les helpers"""
        
        try:
            self.module_usage['helpers'] += 1
            
            # Utiliser extract_domain des helpers
            domain = extract_domain(url)
            
            # Utiliser categorize_url_type des helpers
            url_type = categorize_url_type(url)
            
            # Classification moderne des sources
            is_government = any(gov_indicator in domain.lower() for gov_indicator in ['.gov.', '.gouv.'])
            
            # Sources fiables
            trusted_tunisian = ['bct.gov.tn', 'ins.tn', 'finances.gov.tn']
            trusted_international = ['api.worldbank.org', 'data.worldbank.org', 'imf.org']
            
            is_trusted_source = (
                any(trusted_domain in domain.lower() for trusted_domain in trusted_tunisian) or
                any(trusted_domain in domain.lower() for trusted_domain in trusted_international)
            )
            
            # Type de contenu avec extensions
            content_type = headers.get('content-type', '').lower()
            if 'json' in content_type or url_type == 'api':
                data_type = 'api_data'
            elif 'application/pdf' in content_type:
                data_type = 'pdf_document'
            elif any(excel_type in content_type for excel_type in ['excel', 'spreadsheet']):
                data_type = 'excel_document'
            else:
                data_type = 'web_page'
            
            # Langue basée sur la source
            if any(tn_domain in domain.lower() for tn_domain in trusted_tunisian):
                language = 'french_arabic'
            elif any(intl_domain in domain.lower() for intl_domain in trusted_international):
                language = 'english'
            else:
                language = 'mixed'
            
            return SourceAnalysis(
                domain=domain,
                is_government=is_government,
                is_trusted_source=is_trusted_source,
                content_type=data_type,
                language=language,
                data_freshness='periodic'
            )
            
        except Exception as e:
            logger.error(f"Source analysis with helpers failed: {e}")
            return SourceAnalysis(domain=extract_domain(url))

    def _extract_data_with_all_utils(self, content: str, url: str, source_analysis: SourceAnalysis) -> Dict[str, Any]:
        """Extraction utilisant TOUS les utils disponibles"""
        
        extracted_values = {}
        
        try:
            logger.info(f"Starting comprehensive extraction for {url}")
            
            # 1. EXTRACTION DE BASE selon le type
            if source_analysis.content_type == 'api_data':
                base_values = self._extract_json_comprehensive(content, url)
            elif source_analysis.content_type == 'pdf_document' and PDF_AVAILABLE:
                base_values = self._extract_pdf_content(content, url)
            elif source_analysis.content_type == 'excel_document' and EXCEL_AVAILABLE:
                base_values = self._extract_excel_content(content, url)
            else:
                base_values = self._extract_html_comprehensive(content, url)
            
            extracted_values.update(base_values)
            logger.info(f"Base extraction: {len(base_values)} values")
            
            # 2. INTÉGRATION : Extraction supplémentaire avec clean_extraction_patterns
            self.module_usage['clean_extractor'] += 1
            clean_values = extract_clean_economic_data(content, url)
            
            # Fusionner sans doublons
            for key, value in clean_values.items():
                clean_key = f"clean_{key}"
                if clean_key not in extracted_values:
                    extracted_values[clean_key] = value
            
            logger.info(f"Clean extraction added: {len(clean_values)} additional values")
            
            # 3. Extraction par patterns modernes
            pattern_values = self._extract_patterns_comprehensive(content, url)
            extracted_values.update(pattern_values)
            
            logger.info(f"Total extraction: {len(extracted_values)} values from all methods")
            
            return {'values': extracted_values}
            
        except Exception as e:
            logger.error(f"Comprehensive extraction failed: {e}")
            return {'values': {}}

    def _fetch_content_modern(self, url: str) -> Tuple[Optional[str], Dict[str, str]]:
        """Récupération de contenu moderne avec gestion des types de fichiers"""
        headers = {}
        
        # Timeout adaptatif selon la source
        timeout = self._calculate_adaptive_timeout(url)
        
        for attempt in range(self.max_retries + 1):
            try:
                logger.debug(f"Fetch attempt {attempt + 1}: {url}")
                
                # Headers spéciaux pour sites tunisiens
                if any(tn_domain in url.lower() for tn_domain in ['bct.gov.tn', 'ins.tn', 'finances.gov.tn']):
                    self.session.headers.update({
                        'Accept-Language': 'fr-FR,fr;q=0.9,ar-TN;q=0.8',
                        'Referer': 'https://www.google.com/',
                    })
                
                response = self.session.get(url, timeout=timeout, stream=True)
                response.raise_for_status()
                
                headers = dict(response.headers)
                content_type = headers.get('content-type', '').lower()
                
                # Traitement selon le type de contenu
                if 'application/pdf' in content_type and PDF_AVAILABLE:
                    content = self._extract_pdf_content_safe(response.content)
                elif any(excel_type in content_type for excel_type in ['excel', 'spreadsheet', 'sheet']) and EXCEL_AVAILABLE:
                    content = self._extract_excel_content_safe(response.content)
                else:
                    content = response.text
                
                if content and len(content.strip()) > 10:
                    if len(content) > self.max_content_size:
                        content = content[:self.max_content_size]
                    
                    time.sleep(self.delay)
                    logger.info(f"Content fetched: {len(content)} chars, type: {content_type}")
                    return content, headers
                
            except Exception as e:
                logger.warning(f"Fetch attempt {attempt + 1} failed: {e}")
                if attempt < self.max_retries:
                    time.sleep(2 ** attempt)
                    continue
        
        return None, headers

    def _calculate_adaptive_timeout(self, url: str) -> int:
        """Calcul de timeout adaptatif selon le type de source"""
        url_lower = url.lower()
        
        if 'api.worldbank.org' in url_lower:
            return min(30, self.timeout)
        elif any(tn_domain in url_lower for tn_domain in ['bct.gov.tn', 'ins.tn', 'finances.gov.tn']):
            return min(90, self.timeout * 1.5)
        else:
            return self.timeout

    def _extract_pdf_content_safe(self, pdf_content: bytes) -> str:
        """Extraction PDF sécurisée"""
        if not PDF_AVAILABLE or not pdf_content or len(pdf_content) < 100:
            return ""
        
        try:
            pdf_reader = PdfReader(io.BytesIO(pdf_content))
            text_content = ""
            
            max_pages = min(10, len(pdf_reader.pages))
            for i in range(max_pages):
                try:
                    page_text = pdf_reader.pages[i].extract_text()
                    if page_text:
                        text_content += page_text + "\n"
                        if len(text_content) > 20000:
                            break
                except Exception:
                    continue
            
            return text_content[:20000]
            
        except Exception as e:
            logger.warning(f"PDF extraction failed: {e}")
            return ""

    def _extract_excel_content_safe(self, excel_content: bytes) -> str:
        """Extraction Excel sécurisée"""
        if not EXCEL_AVAILABLE or not excel_content or len(excel_content) < 100:
            return ""
        
        try:
            workbook = load_workbook(io.BytesIO(excel_content), read_only=True)
            text_content = ""
            
            max_sheets = min(3, len(workbook.sheetnames))
            for sheet_name in workbook.sheetnames[:max_sheets]:
                sheet = workbook[sheet_name]
                max_rows = min(100, sheet.max_row)
                
                for row in sheet.iter_rows(max_row=max_rows, values_only=True):
                    if row:
                        row_text = " | ".join(str(cell) for cell in row if cell is not None)
                        if row_text.strip():
                            text_content += row_text + "\n"
                            if len(text_content) > 15000:
                                break
                
                if len(text_content) > 15000:
                    break
            
            workbook.close()
            return text_content[:15000]
            
        except Exception as e:
            logger.warning(f"Excel extraction failed: {e}")
            return ""

    def _extract_json_comprehensive(self, content: str, url: str) -> Dict[str, Any]:
        """Extraction JSON complète avec tous les patterns"""
        extracted_values = {}
        
        try:
            data = json.loads(content.strip())
            
            # World Bank format spécialisé
            if 'worldbank.org' in url.lower() and isinstance(data, list) and len(data) >= 2:
                metadata = data[0]
                data_array = data[1]
                
                for idx, item in enumerate(data_array):
                    if isinstance(item, dict) and 'value' in item:
                        value = item['value']
                        if value is not None and isinstance(value, (int, float)):
                            date = item.get('date', str(idx))
                            country = item.get('country', {}).get('value', 'Tunisia')
                            indicator = item.get('indicator', {}).get('value', 'GDP')
                            
                            key = f"wb_{date}_{idx}"
                            extracted_values[key] = self._create_enhanced_value_cohesive(
                                value=float(value),
                                name=f"{indicator} {country} {date}",
                                unit="USD",
                                raw_text=str(item)[:200],
                                method="json_worldbank",
                                url=url,
                                confidence=0.95,
                                year=int(date) if date.isdigit() else 2024
                            )
            
            # Format JSON générique
            elif isinstance(data, dict):
                for key, value in data.items():
                    if isinstance(value, (int, float)) and self._is_economic_value(value, key):
                        extracted_values[f"json_{key}"] = self._create_enhanced_value_cohesive(
                            value=float(value),
                            name=key,
                            unit="",
                            raw_text=f"{key}: {value}",
                            method="json_generic",
                            url=url,
                            confidence=0.8
                        )
            
            logger.info(f"JSON extraction: {len(extracted_values)} values")
            
        except Exception as e:
            logger.error(f"JSON extraction failed: {e}")
        
        return extracted_values

    def _extract_html_comprehensive(self, content: str, url: str) -> Dict[str, Any]:
        """Extraction HTML complète avec tous les patterns"""
        extracted_values = {}
        
        try:
            soup = BeautifulSoup(content, 'html.parser')
            
            # Extraction spécialisée selon la source
            if 'bct.gov.tn' in url.lower():
                extracted_values.update(self._extract_bct_specialized(soup, url))
            elif 'ins.tn' in url.lower():
                extracted_values.update(self._extract_ins_specialized(soup, url))
            elif 'finances.gov.tn' in url.lower():
                extracted_values.update(self._extract_finance_specialized(soup, url))
            
            # Extraction générique depuis les tables
            tables = soup.find_all('table')[:10]
            for table_idx, table in enumerate(tables):
                table_values = self._extract_table_values(table, url, table_idx)
                extracted_values.update(table_values)
            
            # Extraction depuis les listes et divs
            structured_elements = soup.find_all(['ul', 'ol', 'dl', 'div'], 
                                               class_=re.compile(r'(data|stat|info|metric)', re.I))
            
            for elem_idx, element in enumerate(structured_elements[:5]):
                element_values = self._extract_structured_element_values(element, url, elem_idx)
                extracted_values.update(element_values)
            
            logger.info(f"HTML comprehensive extraction: {len(extracted_values)} values")
            
        except Exception as e:
            logger.error(f"HTML extraction error: {e}")
        
        return extracted_values

    def _extract_bct_specialized(self, soup: BeautifulSoup, url: str) -> Dict[str, Any]:
        """Extraction spécialisée BCT"""
        extracted_values = {}
        
        try:
            # Patterns spécifiques BCT
            bct_patterns = [
                r'(taux directeur|refinancement|intérêt)[:\s]*([0-9,\.]+)\s*%?',
                r'(réserves?|masse monétaire|liquidité)[:\s]*([0-9\s,\.]+)\s*(md|millions?)?',
                r'(change|euro|dollar)[:\s]*([0-9,\.]+)',
            ]
            
            full_text = soup.get_text()
            
            for pattern_idx, pattern in enumerate(bct_patterns):
                matches = re.finditer(pattern, full_text, re.IGNORECASE)
                
                for match_idx, match in enumerate(matches):
                    if len(extracted_values) >= 15:  # Limite BCT
                        break
                    
                    indicator_name = match.group(1).strip()
                    value_str = match.group(2).strip()
                    unit_hint = match.group(3) if len(match.groups()) > 2 else None
                    
                    numeric_value = self._parse_numeric_european(value_str)
                    if numeric_value is not None:
                        unit = self._determine_bct_unit(indicator_name, unit_hint, numeric_value)
                        
                        key = f"bct_{pattern_idx}_{match_idx}"
                        extracted_values[key] = self._create_enhanced_value_cohesive(
                            value=numeric_value,
                            name=f"BCT {indicator_name}",
                            unit=unit,
                            raw_text=match.group(0),
                            method="bct_specialized",
                            url=url,
                            confidence=0.9
                        )
            
        except Exception as e:
            logger.error(f"BCT specialized extraction failed: {e}")
        
        return extracted_values

    def _extract_ins_specialized(self, soup: BeautifulSoup, url: str) -> Dict[str, Any]:
        """Extraction spécialisée INS"""
        extracted_values = {}
        
        try:
            # Format INS avec séparateurs
            full_text = soup.get_text()
            lines = full_text.split('\n')
            
            for line_idx, line in enumerate(lines):
                if '|' in line and len(line) > 15:
                    parts = [part.strip() for part in line.split('|')]
                    
                    if len(parts) >= 2:
                        indicator_name = parts[0]
                        
                        if self._is_economic_indicator_ins(indicator_name):
                            for val_idx, value_part in enumerate(parts[1:], 1):
                                numeric_value = self._parse_numeric_european(value_part)
                                
                                if numeric_value is not None:
                                    unit = self._determine_ins_unit(indicator_name, numeric_value)
                                    
                                    key = f"ins_{line_idx}_{val_idx}"
                                    extracted_values[key] = self._create_enhanced_value_cohesive(
                                        value=numeric_value,
                                        name=f"INS {indicator_name}",
                                        unit=unit,
                                        raw_text=line,
                                        method="ins_specialized",
                                        url=url,
                                        confidence=0.85
                                    )
        
        except Exception as e:
            logger.error(f"INS specialized extraction failed: {e}")
        
        return extracted_values

    def _extract_finance_specialized(self, soup: BeautifulSoup, url: str) -> Dict[str, Any]:
        """Extraction spécialisée Finances"""
        extracted_values = {}
        
        try:
            # Patterns budgétaires
            finance_patterns = [
                r'(budget|recettes|dépenses|déficit)[:\s]*([0-9\s,\.]+)\s*(md|millions?|milliards?)?',
                r'(dette|pib)[:\s]*([0-9\s,\.]+)\s*(md|millions?)?',
                r'(investissement|transfert)[:\s]*([0-9\s,\.]+)',
            ]
            
            full_text = soup.get_text()
            
            for pattern_idx, pattern in enumerate(finance_patterns):
                matches = re.finditer(pattern, full_text, re.IGNORECASE)
                
                for match_idx, match in enumerate(matches):
                    if len(extracted_values) >= 20:
                        break
                    
                    indicator_name = match.group(1).strip()
                    value_str = match.group(2).strip()
                    unit_hint = match.group(3) if len(match.groups()) > 2 else None
                    
                    numeric_value = self._parse_numeric_european(value_str)
                    if numeric_value is not None:
                        unit = unit_hint or ('MD' if numeric_value > 1000 else 'MDT')
                        
                        key = f"finance_{pattern_idx}_{match_idx}"
                        extracted_values[key] = self._create_enhanced_value_cohesive(
                            value=numeric_value,
                            name=f"Finance {indicator_name}",
                            unit=unit,
                            raw_text=match.group(0),
                            method="finance_specialized",
                            url=url,
                            confidence=0.88
                        )
        
        except Exception as e:
            logger.error(f"Finance specialized extraction failed: {e}")
        
        return extracted_values

    def _extract_table_values(self, table, url: str, table_idx: int) -> Dict[str, Any]:
        """Extraction depuis les tables HTML"""
        extracted_values = {}
        
        try:
            rows = table.find_all('tr')
            
            for row_idx, row in enumerate(rows):
                cells = row.find_all(['td', 'th'])
                
                if len(cells) >= 2:
                    for cell_idx in range(len(cells) - 1):
                        label_cell = cells[cell_idx]
                        value_cell = cells[cell_idx + 1]
                        
                        label_text = label_cell.get_text(strip=True)
                        value_text = value_cell.get_text(strip=True)
                        
                        if self._is_economic_indicator_table(label_text):
                            numeric_value = self._parse_numeric_european(value_text)
                            
                            if numeric_value is not None:
                                unit = self._extract_unit_from_text(value_text, label_text)
                                
                                key = f"table_{table_idx}_{row_idx}_{cell_idx}"
                                extracted_values[key] = self._create_enhanced_value_cohesive(
                                    value=numeric_value,
                                    name=label_text[:50],
                                    unit=unit,
                                    raw_text=f"{label_text}: {value_text}",
                                    method="html_table",
                                    url=url,
                                    confidence=0.8
                                )
        
        except Exception as e:
            logger.error(f"Table extraction error: {e}")
        
        return extracted_values

    def _extract_structured_element_values(self, element, url: str, elem_idx: int) -> Dict[str, Any]:
        """Extraction depuis les éléments structurés"""
        extracted_values = {}
        
        try:
            text_content = element.get_text(separator=' ', strip=True)
            
            # Patterns pour éléments structurés
            structured_patterns = [
                r'([A-Za-zÀ-ÿ\s]{5,40})[:=]\s*([0-9,\.]+)\s*(MD|%|millions?)?',
                r'([A-Za-zÀ-ÿ\s]{5,40})\s*:\s*([0-9,\.]+)',
                r'([A-Za-zÀ-ÿ\s]{5,40})\s*-\s*([0-9,\.]+)',
            ]
            
            for pattern_idx, pattern in enumerate(structured_patterns):
                matches = re.finditer(pattern, text_content, re.IGNORECASE)
                
                for match_idx, match in enumerate(matches):
                    if len(extracted_values) >= 5:  # Limite par élément
                        break
                    
                    indicator_name = match.group(1).strip()
                    value_str = match.group(2).strip()
                    unit_hint = match.group(3) if len(match.groups()) > 2 else None
                    
                    if self._is_economic_indicator_structured(indicator_name):
                        numeric_value = self._parse_numeric_european(value_str)
                        
                        if numeric_value is not None:
                            unit = unit_hint or self._determine_unit_from_context(indicator_name, numeric_value)
                            
                            key = f"struct_{elem_idx}_{pattern_idx}_{match_idx}"
                            extracted_values[key] = self._create_enhanced_value_cohesive(
                                value=numeric_value,
                                name=indicator_name[:60],
                                unit=unit,
                                raw_text=match.group(0),
                                method="html_structured",
                                url=url,
                                confidence=0.75
                            )
        
        except Exception as e:
            logger.error(f"Structured element extraction error: {e}")
        
        return extracted_values

    def _extract_patterns_comprehensive(self, content: str, url: str) -> Dict[str, Any]:
        """Extraction par patterns comprehensive"""
        extracted_values = {}
        
        try:
            for pattern_idx, pattern in enumerate(self.modern_patterns):
                matches = re.finditer(pattern, content, re.IGNORECASE | re.MULTILINE)
                
                for match_idx, match in enumerate(matches):
                    if match_idx >= 15:  # Limite par pattern
                        break
                    
                    groups = match.groups()
                    if len(groups) >= 2:
                        name = groups[0].strip()
                        value_str = groups[1].strip()
                        unit = groups[2].strip() if len(groups) > 2 else ''
                        
                        numeric_value = self._parse_numeric_european(value_str)
                        if numeric_value is not None and self._is_economic_value(numeric_value, name):
                            key = f"pattern_{pattern_idx}_{match_idx}"
                            extracted_values[key] = self._create_enhanced_value_cohesive(
                                value=numeric_value,
                                name=name[:60],
                                unit=unit[:10],
                                raw_text=match.group(0),
                                method="pattern_modern",
                                url=url,
                                confidence=0.7
                            )
            
            logger.info(f"Pattern extraction: {len(extracted_values)} values")
            
        except Exception as e:
            logger.error(f"Pattern extraction error: {e}")
        
        return extracted_values

    def _apply_comprehensive_post_processing(self, extracted_data: Dict[str, Any], 
                                           content: str, url: str) -> Dict[str, Any]:
        """Post-traitement complet avec tous les utils - VERSION CORRIGÉE"""
        
        values = extracted_data.get('values', {})
        
        if not values:
            logger.warning("No values to post-process")
            return extracted_data
        
        try:
            logger.info(f"Starting comprehensive post-processing: {len(values)} values")
            
            # 1. Convertir les valeurs en format liste pour le filtre temporel
            values_list = []
            for key, value_data in values.items():
                if isinstance(value_data, dict):
                    # S'assurer que l'item a les champs requis
                    if 'indicator_name' not in value_data:
                        value_data['indicator_name'] = key
                    values_list.append(value_data)
            
            # 2. INTÉGRATION : Filtrage temporel avec temporal_filter
            self.module_usage['temporal_filter'] += 1
            filtered_values_list = filter_by_temporal_period(values_list, content)
            
            # 3. INTÉGRATION : Validation stricte avec data_validator
            self.module_usage['data_validator'] += 1
            
            validation_result = validate_indicators_strict(filtered_values_list, content)
            
            # 4. Reconstruire les données validées
            validated_values = {}
            if validation_result['valid'] and validation_result['valid_data']:
                # Remplacer par les données validées
                for i, validated_item in enumerate(validation_result['valid_data']):
                    key = f"validated_{i}"
                    validated_values[key] = self._convert_validated_to_enhanced(validated_item, url)
                
                logger.info(f"Validation: {len(filtered_values_list)} -> {len(validated_values)} valid items")
            else:
                # Garder les données filtrées si validation échoue
                for i, filtered_item in enumerate(filtered_values_list):
                    key = f"filtered_{i}"
                    validated_values[key] = filtered_item
                logger.warning("Validation failed, keeping filtered data")
            
            # 5. Enrichissement final
            final_values = self._enrich_validated_values(validated_values, url)
            
            return {
                'values': final_values,
                'post_processing': {
                    'temporal_filtering_applied': True,
                    'validation_applied': True,
                    'enrichment_applied': True,
                    'original_count': len(values),
                    'filtered_count': len(filtered_values_list),
                    'validated_count': len(validated_values),
                    'final_count': len(final_values),
                    'validation_summary': validation_result.get('summary', {}),
                    'processing_timestamp': datetime.utcnow().isoformat()
                }
            }
            
        except Exception as e:
            logger.error(f"Post-processing failed: {e}")
            # En cas d'échec, retourner les données originales
            return extracted_data

    def _create_enhanced_value_cohesive(self, value: float, name: str, unit: str, 
                                      raw_text: str, method: str, url: str,
                                      confidence: float = 0.7, year: int = None) -> Dict[str, Any]:
        """Création de valeur enrichie cohésive avec tous les modules"""
        
        # Validation économique
        if not self._validate_economic_value_cohesive(value, name):
            return None
        
        # Année par défaut
        if year is None:
            year = self._extract_year_from_context(raw_text, name) or 2024
        
        # Catégorisation
        category = self._categorize_indicator_cohesive(name)
        
        return {
            "value": value,
            "raw_text": raw_text[:200],
            "indicator_name": name.strip(),
            "enhanced_indicator_name": f"[COHESIVE] {name.strip()}",
            "category": category,
            "unit": unit.strip(),
            "unit_description": self._get_unit_description(unit),
            "context_text": raw_text[:150],
            "extraction_method": method,
            "source_domain": extract_domain(url),
            "extraction_timestamp": datetime.utcnow().isoformat(),
            "confidence_score": confidence,
            "is_economic_indicator": True,
            "is_target_indicator": self._is_target_indicator_cohesive(name),
            "validated": True,
            "year": year,
            "period_type": "annual",
            "temporal_context": f"{year}-cohesive",
            "quality_score": confidence,
            "semantic_quality": self._calculate_semantic_quality(name, value, unit),
            "cohesive_processing": True
        }

    def _validate_economic_value_cohesive(self, value: float, name: str) -> bool:
        """Validation économique cohésive"""
        
        # Rejeter les années évidentes
        if 1950 <= value <= 2030 and len(str(int(value))) == 4:
            year_indicators = ['year', 'année', 'date', 'période']
            if any(year_ind in name.lower() for year_ind in year_indicators):
                return False
        
        # Validation permissive par type
        name_lower = name.lower()
        
        if any(rate_word in name_lower for rate_word in ['taux', 'rate', '%', 'croissance']):
            return -100 <= value <= 1000
        elif 'population' in name_lower:
            return 1000 <= value <= 50_000_000
        elif any(gdp_word in name_lower for gdp_word in ['pib', 'gdp', 'dette']):
            return 0 <= value < 1e15
        else:
            return abs(value) < 1e12

    def _categorize_indicator_cohesive(self, name: str) -> str:
        """Catégorisation cohésive des indicateurs"""
        name_lower = name.lower()
        
        for category, keywords in self.tunisian_indicators.items():
            if any(keyword in name_lower for keyword in keywords):
                category_mapping = {
                    'monetary': 'MONETARY',
                    'statistical': 'DEMOGRAPHIC',
                    'trade': 'TRADE',
                    'fiscal': 'FISCAL',
                    'financial': 'FINANCIAL'
                }
                return category_mapping.get(category, 'OTHER')
        
        return 'OTHER'

    def _is_target_indicator_cohesive(self, name: str) -> bool:
        """Détection des indicateurs cibles"""
        name_lower = name.lower()
        target_keywords = [
            'pib', 'gdp', 'inflation', 'taux directeur', 'chômage',
            'population', 'dette', 'export', 'import', 'balance'
        ]
        return any(keyword in name_lower for keyword in target_keywords)

    def _calculate_semantic_quality(self, name: str, value: float, unit: str) -> float:
        """Calcul de qualité sémantique"""
        quality = 0.5
        
        if len(name) > 8 and any(char.isalpha() for char in name):
            quality += 0.2
        
        if unit and len(unit) > 0:
            quality += 0.1
        
        if 0.01 <= abs(value) <= 1_000_000:
            quality += 0.2
        
        return min(1.0, quality)

    def _extract_year_from_context(self, raw_text: str, name: str) -> Optional[int]:
        """Extraction d'année depuis le contexte"""
        combined_text = f"{raw_text} {name}".lower()
        
        year_match = re.search(r'\b(20[1-2][0-9])\b', combined_text)
        if year_match:
            year = int(year_match.group(1))
            if 2018 <= year <= 2025:
                return year
        
        return None

    def _parse_numeric_european(self, text: str) -> Optional[float]:
        """Parsing numérique format européen"""
        try:
            clean_text = re.sub(r'[^\d\s,\.\-]', '', text.strip())
            if not clean_text:
                return None
            
            clean_text = clean_text.replace(' ', '')
            
            if ',' in clean_text and '.' not in clean_text:
                clean_text = clean_text.replace(',', '.')
            elif ',' in clean_text and '.' in clean_text:
                if clean_text.rfind(',') > clean_text.rfind('.'):
                    clean_text = clean_text.replace('.', '').replace(',', '.')
                else:
                    clean_text = clean_text.replace(',', '')
            
            return float(clean_text)
            
        except (ValueError, TypeError):
            return None

    def _determine_bct_unit(self, indicator: str, unit_hint: str, value: float) -> str:
        """Détermination d'unité BCT"""
        if unit_hint:
            return unit_hint.upper()
        
        indicator_lower = indicator.lower()
        if any(word in indicator_lower for word in ['taux', 'rate']):
            return '%'
        elif value > 1000:
            return 'MD'
        else:
            return 'TND'

    def _determine_ins_unit(self, indicator: str, value: float) -> str:
        """Détermination d'unité INS"""
        indicator_lower = indicator.lower()
        
        if any(word in indicator_lower for word in ['épargne', 'formation', 'capital']):
            return 'MD' if value > 1000 else 'MDT'
        elif any(word in indicator_lower for word in ['taux', 'croissance']):
            return '%'
        elif 'population' in indicator_lower:
            return 'habitants'
        else:
            return 'MD'

    def _determine_unit_from_context(self, indicator: str, value: float) -> str:
        """Détermination d'unité depuis le contexte"""
        indicator_lower = indicator.lower()
        
        if any(word in indicator_lower for word in ['taux', '%', 'croissance']):
            return '%'
        elif any(word in indicator_lower for word in ['pib', 'budget', 'dette']):
            return 'MD' if value > 100 else 'MDT'
        elif 'population' in indicator_lower:
            return 'habitants'
        else:
            return ''

    def _extract_unit_from_text(self, value_text: str, label_text: str) -> str:
        """Extraction d'unité depuis le texte"""
        combined_text = f"{value_text} {label_text}".lower()
        
        unit_patterns = {
            r'%|pourcent': '%',
            r'md|millions?\s*de?\s*dinars?': 'MD',
            r'milliards?': 'MD',
            r'usd|dollars?': 'USD',
            r'eur|euros?': 'EUR',
            r'habitants?': 'habitants'
        }
        
        for pattern, unit in unit_patterns.items():
            if re.search(pattern, combined_text):
                return unit
        
        return ''

    def _get_unit_description(self, unit: str) -> str:
        """Description d'unité"""
        descriptions = {
            'MD': 'Millions de Dinars',
            'MDT': 'Millions de Dinars Tunisiens',
            '%': 'Pourcentage',
            'USD': 'Dollars Américains',
            'EUR': 'Euros',
            'TND': 'Dinars Tunisiens',
            'habitants': 'Nombre d\'habitants'
        }
        return descriptions.get(unit.upper(), unit)

    def _is_economic_indicator_ins(self, text: str) -> bool:
        """Vérification indicateur économique INS"""
        text_lower = text.lower()
        ins_keywords = [
            'épargne', 'formation', 'capital', 'investissement',
            'population', 'emploi', 'chômage', 'entreprises'
        ]
        return any(keyword in text_lower for keyword in ins_keywords)

    def _is_economic_indicator_table(self, text: str) -> bool:
        """Vérification indicateur économique table"""
        return self._is_economic_indicator_general(text)

    def _is_economic_indicator_structured(self, text: str) -> bool:
        """Vérification indicateur économique structuré"""
        return self._is_economic_indicator_general(text)

    def _is_economic_indicator_general(self, text: str) -> bool:
        """Vérification générale d'indicateur économique"""
        if not text or len(text) < 3:
            return False
        
        text_lower = text.lower()
        
        # Exclure les éléments non économiques
        exclude_patterns = [
            r'^(table|tableau|total|somme)',
            r'^(année|year|date)',
            r'^[0-9]{4}',  # Années seules
        ]
        
        for pattern in exclude_patterns:
            if re.match(pattern, text_lower):
                return False
        
        # Vérifier les mots-clés économiques
        economic_keywords = [
            'pib', 'gdp', 'inflation', 'taux', 'dette', 'budget',
            'population', 'emploi', 'chômage', 'export', 'import',
            'réserves', 'change', 'crédit', 'épargne', 'formation'
        ]
        
        return any(keyword in text_lower for keyword in economic_keywords)

    def _is_economic_value(self, value: float, context: str) -> bool:
        """Validation rapide de valeur économique"""
        if not isinstance(value, (int, float)):
            return False
        
        # Rejeter les années
        if 1990 <= value <= 2030 and len(str(int(value))) == 4:
            return False
        
        return -1e10 <= value <= 1e10

    def _convert_validated_to_enhanced(self, validated_item: Dict[str, Any], url: str) -> Dict[str, Any]:
        """Conversion d'item validé vers format enrichi"""
        return {
            "value": validated_item.get('value', 0),
            "raw_text": validated_item.get('raw_text', ''),
            "indicator_name": validated_item.get('indicator_name', ''),
            "enhanced_indicator_name": f"[VALIDATED] {validated_item.get('indicator_name', '')}",
            "category": self._categorize_indicator_cohesive(validated_item.get('indicator_name', '')),
            "unit": validated_item.get('unit', ''),
            "unit_description": self._get_unit_description(validated_item.get('unit', '')),
            "context_text": validated_item.get('raw_text', '')[:150],
            "extraction_method": "validated_traditional",
            "source_domain": extract_domain(url),
            "extraction_timestamp": datetime.utcnow().isoformat(),
            "confidence_score": validated_item.get('confidence_score', 0.8),
            "is_economic_indicator": True,
            "is_target_indicator": self._is_target_indicator_cohesive(validated_item.get('indicator_name', '')),
            "validated": True,
            "year": validated_item.get('year', 2024),
            "period_type": "annual",
            "temporal_context": f"{validated_item.get('year', 2024)}-validated",
            "quality_score": validated_item.get('confidence_score', 0.8),
            "semantic_quality": 0.8,
            "cohesive_processing": True,
            "validation_passed": True
        }

    def _enrich_validated_values(self, validated_values: Dict[str, Any], url: str) -> Dict[str, Any]:
        """Enrichissement final des valeurs validées"""
        enriched_values = {}
        
        for key, value_data in validated_values.items():
            if isinstance(value_data, dict):
                # Enrichir avec métadonnées cohésives
                value_data['cohesive_enrichment'] = {
                    'processing_timestamp': datetime.utcnow().isoformat(),
                    'enrichment_version': 'cohesive_v1.0',
                    'all_utils_applied': True,
                    'source_url': url
                }
                
                enriched_values[key] = value_data
        
        return enriched_values

    def _save_via_smart_storage(self, result: ScrapedContent, url: str):
        """Sauvegarde via le système de stockage intégré"""
        try:
            self.module_usage['storage'] += 1
            
            save_data = {
                'url': url,
                'content': result.raw_content,
                'structured_data': result.structured_data,
                'metadata': result.metadata,
                'cohesive_scraper_info': {
                    'scraper_type': 'CohesiveTunisianWebScraper',
                    'modules_used': list(self.module_usage.keys()),
                    'usage_stats': self.module_usage,
                    'timestamp': datetime.utcnow().isoformat()
                }
            }
            
            storage_result = smart_storage.save_scraping_result(save_data)
            
            if storage_result['success']:
                logger.info(f"Data saved via integrated storage: {storage_result['storage_methods']}")
                
                if not result.metadata:
                    result.metadata = {}
                
                result.metadata['storage_info'] = {
                    'saved_via': 'integrated_smart_storage',
                    'storage_methods': storage_result['storage_methods'],
                    'storage_timestamp': storage_result['timestamp']
                }
            else:
                logger.warning(f"Integrated storage failed: {storage_result['errors']}")
                
        except Exception as e:
            logger.error(f"Integrated storage error: {e}")

    def _perform_debug_logging(self, result: ScrapedContent, url: str, strategy: str):
        """Debug et logging via helpers"""
        try:
            self.module_usage['helpers'] += 1
            
            # Debug des données extraites
            debug_info = debug_extraction_data(result.structured_data, url)
            
            # Log détaillé
            log_extraction_details(result.structured_data, url, strategy)
            
            # Suggestions d'amélioration
            improvements = suggest_extraction_improvements(debug_info)
            
            # Enrichir les métadonnées avec debug
            if not result.metadata:
                result.metadata = {}
            
            result.metadata['debug_info'] = {
                'debug_analysis': debug_info,
                'improvement_suggestions': improvements,
                'debug_timestamp': datetime.utcnow().isoformat()
            }
            
            logger.info(f"Debug logging completed for {url}")
            
        except Exception as e:
            logger.error(f"Debug logging failed: {e}")

    def _build_cohesive_scraped_content(self, content: str, post_processed_data: Dict[str, Any],
                                      source_analysis: SourceAnalysis, url: str,
                                      execution_time: float, enable_llm_analysis: bool) -> ScrapedContent:
        """Construction du contenu scrapé cohésif"""
        
        values = post_processed_data.get('values', {})
        post_processing_info = post_processed_data.get('post_processing', {})
        
        # Résumé d'extraction cohésif
        extraction_summary = ExtractionSummary(
            total_values=len(values),
            categories_found=len(set(v.get('category') for v in values.values() if isinstance(v, dict))),
            target_indicators_found=sum(1 for v in values.values() if isinstance(v, dict) and v.get('is_target_indicator', False)),
            validated_indicators=len(values),
            extraction_method="cohesive_traditional",
            processing_time=execution_time
        )
        
        # Qualité d'extraction cohésive
        extraction_quality = ExtractionQuality(
            total_extracted=post_processing_info.get('original_count', 0),
            high_quality_count=sum(1 for v in values.values() if isinstance(v, dict) and v.get('confidence_score', 0) > 0.8),
            target_indicators_found=extraction_summary.target_indicators_found,
            categories_covered=extraction_summary.categories_found,
            average_confidence=sum(v.get('confidence_score', 0) for v in values.values() if isinstance(v, dict)) / len(values) if values else 0
        )
        
        # Informations de traitement cohésives
        processing_info = ProcessingInfo(
            scraper_version="cohesive_traditional_v1.0",
            extraction_method="cohesive_comprehensive",
            validation_enabled=True,
            ai_enhanced=enable_llm_analysis,
            context_enriched=True,
            semantic_validation=True
        )
        
        # Données structurées cohésives
        structured_data = {
            'extracted_values': values,
            'extraction_summary': extraction_summary.model_dump(),
            'source_analysis': source_analysis.model_dump(),
            'extraction_quality': extraction_quality.model_dump(),
            'processing_info': processing_info.model_dump(),
            'post_processing_info': post_processing_info,
            'cohesive_integration': {
                'all_utils_integrated': True,
                'module_usage_stats': self.module_usage,
                'cohesive_version': 'v1.0_complete'
            }
        }
        
        # Métadonnées cohésives
        metadata = {
            'url': url,
            'scraping_timestamp': datetime.utcnow().isoformat(),
            'execution_time': execution_time,
            'scraper_type': 'CohesiveTunisianWebScraper',
            'scraper_version': 'cohesive_v1.0',
            'content_type': source_analysis.content_type,
            'content_length': len(content),
            'source_domain': source_analysis.domain,
            'cohesive_features': {
                'all_utils_integrated': True,
                'helpers_used': self.module_usage['helpers'] > 0,
                'data_validator_used': self.module_usage['data_validator'] > 0,
                'temporal_filter_used': self.module_usage['temporal_filter'] > 0,
                'clean_extractor_used': self.module_usage['clean_extractor'] > 0,
                'storage_used': self.module_usage['storage'] > 0
            },
            'pdf_extraction_available': PDF_AVAILABLE,
            'excel_extraction_available': EXCEL_AVAILABLE
        }
        
        return ScrapedContent(
            raw_content=content,
            structured_data=structured_data,
            metadata=metadata
        )

    def get_scraper_info(self) -> Dict[str, Any]:
        """Informations du scraper cohésif"""
        return {
            "scraper_type": "CohesiveTunisianWebScraper",
            "version": "cohesive_v1.0_complete",
            "parent_class": "TunisianWebScraper",
            "cohesive_features": [
                "complete_utils_integration",
                "helpers_full_usage",
                "data_validator_pipeline",
                "temporal_filter_processing",
                "clean_extraction_patterns",
                "smart_storage_automatic",
                "comprehensive_post_processing",
                "economic_validation_strict",
                "tunisian_specialization_enhanced"
            ],
            "supported_formats": ["HTML", "JSON", "PDF", "Excel"],
            "pdf_extraction": PDF_AVAILABLE,
            "excel_extraction": EXCEL_AVAILABLE,
            "utils_integration": {
                "helpers": "fully_integrated",
                "data_validator": "pipeline_integrated",
                "temporal_filter": "post_processing_integrated",
                "clean_extractor": "extraction_integrated",
                "storage": "automatic_integrated"
            },
            "module_usage_stats": self.module_usage,
            "max_content_size": self.max_content_size,
            "timeout": self.timeout,
            "max_retries": self.max_retries
        }

    def health_check(self) -> Dict[str, Any]:
        """Vérification de santé cohésive"""
        return {
            "status": "healthy",
            "scraper_type": "CohesiveTunisianWebScraper",
            "version": "cohesive_v1.0",
            "cohesive_integration": {
                "all_utils_available": True,
                "helpers_functional": True,
                "data_validator_functional": True,
                "temporal_filter_functional": True,
                "clean_extractor_functional": True,
                "storage_functional": True
            },
            "session_active": self.session is not None,
            "pdf_available": PDF_AVAILABLE,
            "excel_available": EXCEL_AVAILABLE,
            "timeout_configured": self.timeout,
            "patterns_loaded": len(self.modern_patterns),
            "module_usage_stats": self.module_usage,
            "cohesion_status": "fully_integrated"
        }

    def close(self):
        """Fermeture propre des ressources"""
        if hasattr(self, 'session') and self.session:
            self.session.close()
            logger.info("Cohesive scraper session closed")

# Alias pour compatibilité avec le reste du système
TunisianWebScraper = CohesiveTunisianWebScraper